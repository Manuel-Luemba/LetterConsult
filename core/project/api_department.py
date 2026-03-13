from io import BytesIO
from django.contrib.auth import get_user_model
from django.http import FileResponse
from ninja import Router, File
from ninja.files import UploadedFile
from rest_framework.decorators import authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404
import pandas as pd
from openpyxl import Workbook

from .schemas_department import DepartmentIn, DepartmentOut, PaginatedDepartmentResponse
from core.erp.models import Department
from ..user.models import User

router = Router(tags=["Department"])
REQUIRED_COLUMNS = ["name", "abbreviation", "manager", "cost_center"]

# -------------------------------------------------------------
# CREATE
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.post("")
def create_department(request, payload: DepartmentIn):
    data = payload.dict()

    manager_id = data.pop("manager", None)
    managers_ids = data.pop("managers", [])
    deputy_ids = data.pop("deputy_managers", [])

    department = Department.objects.create(**data)

    if manager_id is not None:
        department.manager = User.objects.get(pk=manager_id)
        if managers_ids is not None and manager_id not in managers_ids:
            managers_ids.append(manager_id)
    department.save()

    if managers_ids:
        department.managers.set(managers_ids)
    if deputy_ids:
        department.deputy_managers.set(deputy_ids)

    return {"success": True, "data": DepartmentOut.from_orm(department)}

# -------------------------------------------------------------
# IMPORT EXCEL
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.post("/import-excel")
def import_departments_excel(request, file: UploadedFile = File(...)):
    try:
        df = pd.read_excel(file.file)
    except Exception as e:
        return {"error": f"Erro ao ler o Excel: {str(e)}"}

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        return {"error": f"Colunas obrigatórias ausentes: {', '.join(missing)}"}

    created = []
    errors = []

    for idx, row in df.iterrows():
        try:
            if pd.isna(row["name"]) or pd.isna(row["cost_center"]):
                errors.append(f"Linha {idx + 2}: faltam dados obrigatórios")
                continue

            is_active = True
            if "is_active" in df.columns and not pd.isna(row.get("is_active")):
                value = str(row["is_active"]).strip().upper()
                is_active = value in ["SIM", "YES", "TRUE", "1", "S", "Y"]

            department = Department.objects.create(
                name=row["name"],
                desc=row.get("desc", ""),
                abbreviation=row.get("abbreviation", None),
                cost_center=row["cost_center"],
                is_active=is_active,
            )

            if not pd.isna(row.get("manager")):
                try:
                    manager_id = int(row["manager"])
                    department.manager_id = manager_id
                    department.save()
                    department.managers.add(manager_id)
                except (ValueError, TypeError):
                    errors.append(f"Linha {idx + 2}: manager_id inválido")

            created.append(department.id)

        except Exception as e:
            errors.append(f"Linha {idx + 2}: {str(e)}")

    return {
        "success": True,
        "imported_ids": created,
        "total_imported": len(created),
        "errors": errors,
        "total_errors": len(errors),
    }

# -------------------------------------------------------------
# DOWNLOAD TEMPLATE
# -------------------------------------------------------------
@router.get("/download-template")
def download_department_template(request):
    User = get_user_model()
    users = User.objects.all().values("id", "first_name", "last_name", "email")
    departments = Department.objects.all().values("id", "abbreviation", "name", "cost_center", "is_active", "manager")

    wb = Workbook()

    ws = wb.active
    ws.title = "Modelo_Departamentos"
    # ws.append(["name", "abbreviation", "manager", "cost_center", "is_active"])
    # ws.append(["Departamento de TI", "TI", 1, "CC-001", "SIM"])
    # ws.append(["Departamento Financeiro", "FIN", 2, "CC-002", "NÃO"])

    ws2 = wb.create_sheet("Instruções")
    ws2.append(["INSTRUÇÕES PARA IMPORTAÇÃO DE DEPARTAMENTOS"])
    ws2.append([""])
    ws2.append(["COLUNAS OBRIGATÓRIAS:"])
    ws2.append(["• name - Nome do departamento"])
    ws2.append(["• abbreviation - Abreviatura única"])

    ws2.append(["• cost_center - Centro de custo"])
    ws2.append([""])
    ws2.append(["COLUNAS OPCIONAIS:"])
    ws2.append(["• is_active - SIM/NÃO, padrão: SIM"])


    ws2.append([""])
    ws.append(["LISTA DE DEPARTAMENTOS EXISTENTES:"])
    ws.append(["Abreviatura", "Nome", "Centro de Custo", "Ativo?"])
    for d in departments:
        ws.append([d["abbreviation"], d["name"], d["cost_center"], d["is_active"]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(output, filename="modelo_importacao_departamentos.xlsx")

@router.get("/download-template1")
def download_department_template1(request):
    User = get_user_model()
    users = list(User.objects.all().values("id", "first_name", "last_name", "email"))
    departments = list(Department.objects.all().values("id", "abbreviation", "name"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo_Departamentos"

    ws.append(["name", "abbreviation", "manager", "cost_center", "is_active"])
    ws.append(["Departamento de TI", "TI", 1, "CC-001", "SIM"])
    ws.append(["Departamento Financeiro", "FIN", 2, "CC-002", "NÃO"])

    ws2 = wb.create_sheet("Instruções")
    ws2.append(["INSTRUÇÕES PARA IMPORTAÇÃO DE DEPARTAMENTOS"])
    ws2.append([""])

    ws2.append(["LISTA DE DEPARTAMENTOS EXISTENTES:"])
    ws2.append(["ID", "Abreviatura", "Nome"])

    for d in departments:
        ws2.append([d["id"], d["abbreviation"], d["name"]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from datetime import datetime
    filename = f"modelo_departamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return FileResponse(output, filename=filename)
# -------------------------------------------------------------
# LIST PAGINATED
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.get("/list", response=PaginatedDepartmentResponse)
def list_departments(request, page: int = 1, page_size: int = 20):
    queryset = Department.objects.all().order_by('-id')
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page)
    items = [DepartmentOut.from_orm(dep) for dep in page_obj]

    return {
        "count": paginator.count,
        "total_pages": paginator.num_pages,
        "current_page": page_obj.number,
        "page_size": page_size,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
        "items": items,
    }

# -------------------------------------------------------------
# GET BY ID
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.get("/{department_id}", response=DepartmentOut)
def get_department(request, department_id: int):
    dep  = get_object_or_404(Department, id=department_id)
    return DepartmentOut.from_orm(dep)

# -------------------------------------------------------------
# ACTIVATE / DEACTIVATE
# -------------------------------------------------------------
@router.patch("/{department_id}/deactivate")
def deactivate_department(request, department_id: int):
    department = get_object_or_404(Department, id=department_id)
    department.is_active = False
    department.save()
    return {"success": True}

@router.patch("/{department_id}/activate")
def activate_department(request, department_id: int):
    department = get_object_or_404(Department, id=department_id)
    department.is_active = True
    department.save()
    return {"success": True}

# -------------------------------------------------------------
# UPDATE
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.put("/{department_id}")
def update_department(request, department_id: int, payload: DepartmentIn):
    department = get_object_or_404(Department, id=department_id)
    data = payload.dict()

    manager_id = data.pop("manager", None)
    managers_ids = data.pop("managers", None)
    deputy_ids = data.pop("deputy_managers", None)

    for attr, value in data.items():
        setattr(department, attr, value)
    department.save()

    if manager_id is not None:
        department.manager = User.objects.get(pk=manager_id)
        department.save()
        if managers_ids is not None and manager_id not in managers_ids:
            managers_ids.append(manager_id)

    if managers_ids is not None:
        department.managers.set(managers_ids)
    if deputy_ids is not None:
        department.deputy_managers.set(deputy_ids)
    return {"success": True}

# -------------------------------------------------------------
# DELETE
# -------------------------------------------------------------
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
@router.delete("/{department_id}")
def delete_department(request, department_id: int):
    department = get_object_or_404(Department, id=department_id)
    department.delete()
    return {"success": True}